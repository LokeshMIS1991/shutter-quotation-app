import io
import json
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pypdfium2 as pdfium  # PDF to Image Rendering
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, Image as RLImage, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import streamlit as st

st.set_page_config(
    page_title="Sidharth Shutter & Automation Pvt. Ltd.",
    layout="wide",
    page_icon="🛡️",
)

# --- PRODUCTS LIST STRUCTURE ---
products = [
    {
        "productItem": "Rolling Shutters",
        "subCategory": "Motorized, Gear & Manual Rolling Shutters",
        "icon": "🌀",
    },
    {
        "productItem": "Dock Leveler",
        "subCategory": "Hydraulic Doclevller, Hydraulic DockEdge, Manual DockEdge",
        "icon": "🏗️",
    },
    {
        "productItem": "Gates",
        "subCategory": "Sliding, Telescopic, L-Folding, Swing & Retractable Gates",
        "icon": "🚪",
    },
    {
        "productItem": "Doors",
        "subCategory": "High Speed, Fire, HMPS, GPD & Overhead Sectional Doors",
        "icon": "🚪",
    },
    {
        "productItem": "Boom Barrier",
        "subCategory": "Automatic & Heavy-Duty Traffic Barriers",
        "icon": "🚧",
    },
    {
        "productItem": "Dock Shelter",
        "subCategory": "Retractable & Inflatable Dock Shelters",
        "icon": "🏬",
    },
    {
        "productItem": "Dock Bumper",
        "subCategory": "Heavy Rubber & Moulded Bumpers",
        "icon": "🛡️",
    },
]

# --- CUSTOM CSS FOR ENHANCED ENTERPRISE LOOK ---
st.markdown(
    """
<style>
    .stApp {
        background-color: #f8fafc;
    }
    
    /* Welcome Header Banner */
    .welcome-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        padding: 45px 30px;
        border-radius: 18px;
        color: white;
        text-align: center;
        margin-bottom: 30px;
        box-shadow: 0 10px 20px -3px rgba(0, 0, 0, 0.15);
    }
    .welcome-header h1 {
        color: #ffffff !important;
        font-weight: 800;
        margin: 0;
        font-size: 36px;
        letter-spacing: 0.5px;
    }
    .welcome-header p {
        color: #94a3b8 !important;
        margin-top: 12px;
        margin-bottom: 0;
        font-size: 18px;
    }
    
    .stButton>button {
        border-radius: 8px;
        font-weight: 600;
    }
</style>
""",
    unsafe_allow_html=True,
)

# --- SESSION STATE FOR NAVIGATION ---
if "selected_product" not in st.session_state:
    st.session_state["selected_product"] = "Home"


def navigate_to(page_name):
    st.session_state["selected_product"] = page_name


# --- PERMANENT SPECIFICATIONS FILE MANAGEMENT ---
SPEC_FILE = "specifications_master.json"

DEFAULT_SPECS = {
    "slat_nat_list": [
        "78mm (H) x 1mm Thick Galvalume Curved Slat in Natural Finish",
        "78mm (H) x .80mm Thick Galvalume Curved Slat in Natural Finish",
        "78mm (H) x .90mm Thick Galvalume Curved Slat in Natural Finish",
        "78mm (H) x 1mm Thick Galvanized in Curved Slat",
        "78mm (H) x .80mm Thick Galvanized in Curved Slat",
        "78mm (H) x .90mm Thick Galvanized in Curved Slat",
        "90mm (H) x .80mm Thick Galvalume plain slats",
        "90mm (H) x .90mm Thick Galvalume plain slats in Natural Finish",
        "90mm (H) x .80mm Thick Galvanized in plain slats",
        "90mm (H) x .90mm Thick Galvanized in plain slats",
        "120mm (H) x .80mm Big Profile Galvanized in plain slats",
        "120mm (H) x .90mm Big Profile Galvanized in plain slats",
        "120mm (H) x .1mm Big Profile Galvanized in plain slats",
        "120mm (H) x 1.2mm Big Profile Galvanized in plain slats",
        "120mm (H) x .80mm Big Profile Galvalume in plain slats",
        "120mm (H) x .90mm Big Profile Galvalume in plain slats",
        "120mm (H) x .1mm Big Profile Galvalume in plain slats",
        "120mm (H) x 1.2mm Big Profile Galvalume in plain slats"
    ],
    "slat_pow_list": [
        "Finish - Powder Coating As Per RAL",
        "Finish - Red Oxide",
        "Finish - PU Paint As Per RAL",
        "Finish - Enamel Paint As Per RAL"
    ],
    "guide_list": [
        "TG Guide with Rubber Seal",
        "TG Guide with Rubber Seal with Grey Epoxy Finish",
        "TG Guide With Grey Epoxy Finish",
        "U Guide with Grey Epoxy Finish",
        "U Guide",
        "U Guide With Rubber"
    ],
    "bottom_list": [
        "Super Bottom with Rubber Seal with Grey Epoxy Finish",
        "Super Bottom with Rubber Seal",
        "Aluminium Bottom with Rubber Seal with Grey Epoxy Finish",
        "Aluminium Bottom with Rubber Seal"
    ],
    "hood_list": [
        ".80mm Thick Galvalume Hood & Motor Cover in Natural Finish",
        ".90mm Thick Galvalume Hood & Motor Cover in Natural Finish",
        "1.2mm Thick Galvalume Hood & Motor Cover in Natural Finish",
        "1mm Thick Galvalume Hood & Motor Cover in Natural Finish",
        ".80mm Thick Galvanized Hood & Motor Cover",
        ".90mm Thick Galvanized Hood & Motor Cover",
        "1mm Thick Galvanized Hood & Motor Cover",
        "1.2mm Thick Galvanized Hood & Motor Cover",
        ".80mm Thick Galvalume Hood Cover in Natural Finish",
        ".80mm Thick Galvanized Hood Cover",
        "N/A"
    ],
    # UPDATED DOCK LEVELER MASTER SPECIFICATION OPTIONS ACCORDING TO LINE ORDER
    "dock_ce_cert_list": [
        "CE certified: European norm 1398",
        "N/A"
    ],
    "dock_capacity_list": [
        "15T Uniform distributed load (UDL) and 6T point load",
        "22T Uniform distributed load (UDL) and 9T point load"
    ],
    "dock_cylinder_list": [
        "Single Lift cylinder",
        "Single Lip cylinder"
    ],
    "dock_height_adj_list": [
        "+640mm/-300mm",
        "N/A"
    ],
    "dock_bumper_list": [
        "Super Bumper",
        "Mega Bumper"
    ]
}


def load_specifications():
    if os.path.exists(SPEC_FILE):
        try:
            with open(SPEC_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                for k, v in DEFAULT_SPECS.items():
                    if k not in data:
                        data[k] = v
                return data
        except Exception:
            return DEFAULT_SPECS.copy()
    return DEFAULT_SPECS.copy()


specs = load_specifications()
for key, value in specs.items():
    if key not in st.session_state:
        st.session_state[key] = value

SAFETY_LOCK_OPTIONS = [
    "Wind Locks",
    "Storm Anchors",
    "External Safety Break",
    "Side Locks-2 Nos",
    "Standard Center Lock with 2 Keys",
    "Gear Set - Both Side",
    "Gear Set - Single Side (Outside)",
]

OPERATOR_OPTIONS = [
    "CE Certified Indirect Drive Brand Strong Life Sidharth Make",
    "Three Station Push Button",
]

DOCK_SAFETY_OPTIONS = [
    "Yellow/Black Hazard Safety Toe Guards",
    "Emergency Stop Hydraulic Velocity Fuse Safety Valve",
    "Maintenance Safety Support Strut Bar",
    "Automatic Wheel Chock Interlock Integration"
]

PRODUCT_HIERARCHY = {
    "Rolling Shutters": [
        "Motorized Rolling Shutter",
        "Gear Rolling Shutter",
        "Manual Rolling Shutter",
    ],
    "Dock Leveler": [
        "Hydraulic Doclevller",
        "Hydraulic DockEdge",
        "Manual DockEdge",
    ],
    "Gates": [
        "Sliding Gate",
        "Telescopic Gate",
        "L-Folding Gate",
        "Swing Gate",
        "Retractable Gate",
    ],
    "Doors": [
        "High Speed Door",
        "Fire Door",
        "HMPS Door",
        "GPD Door",
        "Overhead Sectional Door",
    ],
    "Boom Barrier": [
        "Automatic Traffic Barrier",
        "Heavy-Duty Traffic Barrier",
    ],
    "Dock Shelter": [
        "Retractable Dock Shelter",
        "Inflatable Dock Shelter",
    ],
    "Dock Bumper": [
        "Heavy Rubber Bumper",
        "Moulded Bumper",
    ]
}

# DEFAULT DATA FOR SECTION A (PAGE 2)
DEFAULT_ABOUT_US_LIST = [
    {
        "title": "1. Core Company Highlights & Industrial Legacy",
        "text": "• Over Two Decades of Engineering Excellence: Established in 1998, Sidharth Shutter & Automation Pvt. Ltd. (SSAPL) has evolved into a premier national leader in high-performance industrial entrance and access automation systems across India.\n• Extensive Nationwide Deployment: Successfully engineered, custom-manufactured, and commissioned over 10,000+ heavy-duty industrial shutters, dock systems, and automated security barriers for leading public and private enterprises.\n• Pan-India Operational Footprint: Serving top-tier logistics parks, manufacturing plants, warehouse hubs, defense facilities, commercial complexes, and infrastructure projects with robust regional project execution capability.\n• Advanced Manufacturing Facility: Powered by state-of-the-art infrastructure featuring CNC roll-forming machinery, high-precision laser cutters, robotic welding units, and fully automated powder-coating lines."
    },
    {
        "title": "2. Engineering, Quality Standards & Certification",
        "text": "• ISO & CE Certified Quality Benchmarks: Strict compliance with ISO 9001:2015 Quality Management Systems and European CE safety certification standards across all product development cycles.\n• High Wind-Load & Extreme Weather Resilience: Structural curtains and frames are specifically calculated and stress-tested to withstand high wind velocity, pressure differentials, severe rain, and harsh industrial environments.\n• Premium Heavy-Gauge Raw Materials: Exclusively using certified high-tensile Galvalume, Galvanized Steel (GI), and structural aluminum alloys for exceptional structural stability, long operational lifespan, and maximum corrosion protection."
    },
    {
        "title": "3. Comprehensive Product & Solutions Portfolio",
        "text": "• Industrial & Commercial Rolling Shutters: Heavy-duty motorized, gear-operated, and manual roll-up systems designed for maximum perimeter security, thermal insulation, and continuous high-frequency operations.\n• High-Speed & Specialized Doors: Rapid roll-up PVC doors, insulated sectional overhead doors, fire-rated HMPS doors, and acoustic/cleanroom barriers engineered for dynamic airflow control.\n• Advanced Entrance Automation: Automatic boom barriers, cantilever sliding gates, telescopic gates, heavy-duty swing gates, and motor-driven retractable security gates.\n• Integrated Loading Bay Equipment: Hydraulic and mechanical dock levelers, inflatable/retractable dock shelters, heavy-duty rubber bumpers, and integrated loading dock safety accessories."
    },
    {
        "title": "4. The Sidharth Advantage: Turnkey Execution & Support",
        "text": "• End-to-End Turnkey Project Management: Complete project lifecycle execution—from initial site survey, structural shop drawings, and custom fabrication to installation, electrical integration, and final handover.\n• 24x7 Rapid Service & Maintenance Network: Highly specialized field service engineering team delivering preventative maintenance, fast emergency troubleshooting, AMC services, and guaranteed original spare parts supply.\n• Architectural & Operational Customization: Tailor-made designs engineered to integrate smoothly into non-standard structural openings, unique headroom spaces, and specialized warehouse management workflows."
    },
    {
        "title": "5. Technical R&D, Testing & Safety Protocols",
        "text": "• In-House Research & Innovation: Continuous R&D focused on ultra-quiet drive mechanisms, high-wind slat locking mechanisms, smart sensor technology, and long-life motor controls.\n• Rigorous Multi-Cycle Factory Testing: Every motor drive, control unit, and shutter curtain undergoes strict pre-dispatch factory testing for smooth mechanical movement, limit accuracy, and electrical load capacity.\n• Uncompromising Operational Safety: Integrated emergency manual hand-chain system, optical safety sensors, bottom-edge safety buffers, and external dynamic drop-brake protection standard across heavy-duty configurations."
    }
]

if "about_us_data" not in st.session_state:
    st.session_state["about_us_data"] = DEFAULT_ABOUT_US_LIST.copy()

# DOCK LEVELER TECH SPECS FOR SECTION C (PAGE 4)
DEFAULT_TECH_SPECS = [
    {"param": "Equipment Type", "spec": "Electro-Hydraulic Station Dock Leveler"},
    {"param": "1. CE Certification", "spec": "CE certified: European norm 1398"},
    {"param": "2. Load Capacity", "spec": "15T Uniform distributed load (UDL) and 6T point load"},
    {"param": "3. Cylinders Assembly", "spec": "Single Lift cylinder / Single Lip cylinder"},
    {"param": "4. Height Adjustment", "spec": "+640mm/-300mm"},
    {"param": "5. Bumper", "spec": "Super Bumper / Mega Bumper"},
    {"param": "Platform Dimensions", "spec": "Width: 2000 mm x Length: 2500 mm (Pit Depth: 600 mm)"},
    {"param": "Control Panel Box", "spec": "Wall-mounted IP65 Control Box with Dead-Man Push Buttons and Emergency Stop Switch"},
    {"param": "Safety Toe Guards", "spec": "Full-length yellow/black hazard striped steel toe guards on both lateral sides"},
    {"param": "Safety Drop Valve", "spec": "Automatic Hydraulic Emergency Velocity Fuse Safety Valve in case of sudden truck departure"},
    {"param": "Operating Temperature", "spec": "-10°C to +55°C Industrial Operating Range"}
]

if "tech_specs_data" not in st.session_state:
    st.session_state["tech_specs_data"] = DEFAULT_TECH_SPECS.copy()

DEFAULT_TERMS = [
    {
        "category": "Unloading & Handling",
        "details": "• Unloading of material at site is strictly in client scope.\n• All required civil pit work, pocket cutting, and masonry work is in client scope.\n• Hydra crane, forklift, and heavy material handling equipment to be arranged by client."
    },
    {
        "category": "Storage & Material Handling",
        "details": "• Safe, dry, and locked storage space for materials at site until installation completion.\n• Material verification against packing list and shifting to exact pit location is client's responsibility."
    },
    {
        "category": "Electrical Scope",
        "details": "• Three-phase 415V AC power supply with dedicated MCB up to installation point by client.\n• Standard socket and free electric power point within 5 meters of installation spot."
    },
    {
        "category": "Warranty Terms",
        "details": "• Automation system is warranted against manufacturing defects for 12 months from installation or 13 months from invoice date, whichever is earlier."
    },
    {
        "category": "Payment & Price Validity",
        "details": "• 75% Advance along with formal Purchase Order.\n• 25% balance payment prior to material dispatch against proforma invoice.\n• This commercial offer remains valid for 20 days from the date of issuance."
    }
]

if "terms_data" not in st.session_state:
    st.session_state["terms_data"] = DEFAULT_TERMS.copy()

# ==============================================================================
# PAGE 1: HOME / LANDING PAGE
# ==============================================================================
if st.session_state["selected_product"] == "Home":

    st.markdown(
        """
    <div class="welcome-header">
        <h1>Welcome To Sidharth Shutter & Automation Pvt. Ltd.</h1>
        <p>Industrial Access & Entrance Automation Solutions</p>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.write("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown(
            """
            <div style="background-color: white; padding: 30px; border-radius: 12px; border: 1px solid #e2e8f0; text-align: center; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);">
                <h3 style="color: #0f172a; margin-bottom: 10px;">Create Professional Quotation</h3>
                <p style="color: #64748b; font-size: 14px; margin-bottom: 25px;">
                    Click below to start generating custom multi-page commercial offers for Rolling Shutters and Dock Levelers.
                </p>
            </div>
            """,
            unsafe_allow_html=True
        )
        st.write("")
        if st.button("🚀 Create Quotation", use_container_width=True, type="primary"):
            st.session_state["selected_product"] = "Rolling Shutters"
            st.rerun()

# ==============================================================================
# PAGE 2: QUOTATION BUILDER (ROLLING SHUTTERS & DOCK LEVELER)
# ==============================================================================
elif st.session_state["selected_product"] in ["Rolling Shutters", "Dock Leveler"]:

    col_nav1, col_nav2 = st.columns([1, 6])
    with col_nav1:
        if st.button("🏠 Home", use_container_width=True):
            st.session_state["selected_product"] = "Home"
            st.rerun()
    with col_nav2:
        st.title("📋 Industrial Products Quotation Builder")

    st.sidebar.header("📋 Client & Proposal Details")

    client_id = st.sidebar.text_input("Client ID", "SSAPL-CL-8842")
    client_name = st.sidebar.text_input(
        "Company Name", "M/S ABHINAV INFRA BUILD PVT. LTD."
    )
    attn_person = st.sidebar.text_input("Attention To / Kind Attn", "Mr. Sharma")
    contact_details = st.sidebar.text_input("Contact Details", "Mob:-8889911529")
    shipping_address = st.sidebar.text_area(
        "Shipping Address",
        "RGLP PITHAMPUR Sector-1 Pithampur MADHYA PRADESH 454774",
        height=70,
    )
    site_person = st.sidebar.text_input("Site Person Mobile", "9926952398")
    billing_address = st.sidebar.text_area(
        "Billing Address",
        "207,208, Industry House, Indore, MP 452001",
        height=70,
    )
    gstin = st.sidebar.text_input("Client GSTIN", "23AAHCA9425D1ZY")

    st.sidebar.markdown("---")
    c_date, c_ref = st.sidebar.columns(2)
    qtn_date = c_date.date_input("Quotation Date")
    qtn_ref_no = c_ref.text_input("Qtn Ref No", "SSAPL/2026-27/319R2")

    sales_reference_1 = st.sidebar.text_input(
        "Sales Reference 1", "Mr. Nishant (90010 42914)"
    )
    sales_reference_2 = st.sidebar.text_input(
        "Sales Reference 2", "Mr. Jeevan Sharma (9828771899)"
    )

    quotation_made_by = st.sidebar.text_input(
        "Quotation Made By", "Lokesh MIS"
    )

    # PAGE SELECTION TOGGLES
    st.sidebar.markdown("---")
    st.sidebar.header("📄 Page Selection Settings")
    include_page_2 = st.sidebar.checkbox("Include Page 2 (About Us)", value=True)
    include_page_4 = st.sidebar.checkbox("Include Page 4 (Technical Specs)", value=True)
    include_page_5 = st.sidebar.checkbox("Include Page 5 (Terms & Conditions)", value=True)

    # EXPANDERS FOR CUSTOMIZATION
    with st.expander("📄 Page 1: Cover Letter Text Customization", expanded=False):
        cover_body_text = st.text_area(
            "Cover Letter Intro Body",
            "Dear Sir / Ma'am,\n\n"
            "We extend our sincere gratitude for the interest you have shown in our products and services. "
            "It is our privilege to present this detailed commercial offer, crafted to precisely address "
            "the loading bay and dock automation requirements of your facility.\n\n"
            "This proposal package consists of the detailed specifications and commercial offer enclosed herewith.\n\n"
            "Warm Regards,",
            height=200
        )

    with st.expander("📄 Page 2: About Us (Company Profile) Text Customization", expanded=False):
        updated_about = []
        for idx, item in enumerate(st.session_state["about_us_data"]):
            col_t, col_d, col_rm = st.columns([2, 3, 1])
            with col_t:
                t_val = st.text_input(f"Section #{idx+1} Title", value=item["title"], key=f"ab_title_{idx}")
            with col_d:
                x_val = st.text_area(f"Section #{idx+1} Details", value=item["text"], height=100, key=f"ab_text_{idx}")
            with col_rm:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("❌ Remove", key=f"rm_ab_{idx}"):
                    st.session_state["about_us_data"].pop(idx)
                    st.rerun()
            updated_about.append({"title": t_val, "text": x_val})
        st.session_state["about_us_data"] = updated_about

    with st.expander("📄 Page 4: Section C - Technical Specifications Customization", expanded=False):
        tech_intro_text = st.text_area(
            "Header Introductory Text",
            "The following specifications define the precise engineering parameters for the Dock Leveler / Industrial Equipment offered.",
            height=70
        )
        updated_specs = []
        for idx, item in enumerate(st.session_state["tech_specs_data"]):
            col_p, col_s, col_rm = st.columns([2, 3, 1])
            with col_p:
                p_val = st.text_input(f"Parameter #{idx+1}", value=item["param"], key=f"tp_{idx}")
            with col_s:
                s_val = st.text_input(f"Specification #{idx+1}", value=item["spec"], key=f"ts_{idx}")
            with col_rm:
                if st.button("❌ Remove", key=f"rm_ts_{idx}"):
                    st.session_state["tech_specs_data"].pop(idx)
                    st.rerun()
            updated_specs.append({"param": p_val, "spec": s_val})
        st.session_state["tech_specs_data"] = updated_specs
        
        if st.button("➕ Add Technical Specification"):
            st.session_state["tech_specs_data"].append({"param": "New Parameter", "spec": "Description"})
            st.rerun()

        tech_note_text = st.text_input("Footer Note Text", "Note: Technical specifications are subject to final site measurements and drawing approval.")

    with st.expander("📄 Page 5: Section D - Terms & Conditions Customization", expanded=False):
        exclusions_subhead = st.text_input("Sub-Header Title", "Exclusions - Client Scope & Operational Terms")
        updated_terms = []
        for idx, term in enumerate(st.session_state["terms_data"]):
            col_tc, col_td, col_rm = st.columns([2, 3, 1])
            with col_tc:
                c_val = st.text_input(f"Category #{idx+1}", value=term["category"], key=f"tc_cat_{idx}")
            with col_td:
                d_val = st.text_area(f"Details #{idx+1}", value=term["details"], height=70, key=f"tc_det_{idx}")
            with col_rm:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("❌ Remove", key=f"rm_tc_{idx}"):
                    st.session_state["terms_data"].pop(idx)
                    st.rerun()
            updated_terms.append({"category": c_val, "details": d_val})
        st.session_state["terms_data"] = updated_terms

    # --- ITEMS MANAGEMENT IN SESSION STATE ---
    st.markdown("### 📦 Product Specifications & Pricing")

    if "shutter_items" not in st.session_state:
        st.session_state["shutter_items"] = [{
            "main_product": "Dock Leveler",
            "type": "Hydraulic Doclevller",
            "dock_ce_cert": ["CE certified: European norm 1398"],
            "dock_capacity": ["15T Uniform distributed load (UDL) and 6T point load"],
            "dock_cylinder": ["Single Lift cylinder"],
            "dock_height_adj": ["+640mm/-300mm"],
            "dock_bumper": ["Super Bumper"],
            "dock_safety": [
                "Yellow/Black Hazard Safety Toe Guards",
                "Emergency Stop Hydraulic Velocity Fuse Safety Valve"
            ],
            "slat_nat": [],
            "slat_pow": [],
            "guide": [],
            "bottom": [],
            "hood": [],
            "safety_locks": [],
            "operator": [],
            "hsn": "84289090",
            "width": 2000,
            "height": 2500,
            "qty": 2,
            "mat_rate": 245000,
            "inst_rate": 12000,
        }]

    def add_product_item():
        st.session_state["shutter_items"].append({
            "main_product": "Dock Leveler",
            "type": "Hydraulic Doclevller",
            "dock_ce_cert": ["CE certified: European norm 1398"],
            "dock_capacity": ["15T Uniform distributed load (UDL) and 6T point load"],
            "dock_cylinder": ["Single Lift cylinder"],
            "dock_height_adj": ["+640mm/-300mm"],
            "dock_bumper": ["Super Bumper"],
            "dock_safety": [],
            "slat_nat": [],
            "slat_pow": [],
            "guide": [],
            "bottom": [],
            "hood": [],
            "safety_locks": [],
            "operator": [],
            "hsn": "84289090",
            "width": 2000,
            "height": 2500,
            "qty": 1,
            "mat_rate": 245000,
            "inst_rate": 12000,
        })

    def remove_product_item(index):
        if len(st.session_state["shutter_items"]) > 1:
            st.session_state["shutter_items"].pop(index)

    for idx, item in enumerate(st.session_state["shutter_items"]):
        selected_sub_cat = item.get("type", "Hydraulic Doclevller")
        with st.expander(f"📌 Item #{idx + 1}: {selected_sub_cat}", expanded=True):
            col_p_main, col_p_cat, col_del = st.columns([2, 2, 1])
            
            with col_p_main:
                curr_main_prod = item.get("main_product", "Dock Leveler")
                main_prod_keys = list(PRODUCT_HIERARCHY.keys())
                main_idx = main_prod_keys.index(curr_main_prod) if curr_main_prod in main_prod_keys else 1
                
                selected_main = st.selectbox(
                    f"Product Item #{idx + 1}",
                    main_prod_keys,
                    index=main_idx,
                    key=f"main_prod_{idx}"
                )
                item["main_product"] = selected_main

            with col_p_cat:
                available_sub_cats = PRODUCT_HIERARCHY.get(selected_main, ["Hydraulic Doclevller"])
                curr_sub = item.get("type", available_sub_cats[0])
                sub_idx = available_sub_cats.index(curr_sub) if curr_sub in available_sub_cats else 0
                
                item["type"] = st.selectbox(
                    f"Category / Sub-Category #{idx + 1}",
                    available_sub_cats,
                    index=sub_idx,
                    key=f"sub_cat_{idx}"
                )

            with col_del:
                st.markdown("<br>", unsafe_allow_html=True)
                if len(st.session_state["shutter_items"]) > 1:
                    st.button("❌ Remove", key=f"del_{idx}", on_click=remove_product_item, args=(idx,))

            item["hsn"] = st.text_input("HSN Code", value=item.get("hsn", "84289090"), key=f"hsn_{idx}")

            # --- DOCK LEVELER SPECIFIC UI FIELDS (UPDATED SEQUENCE 1 TO 5) ---
            if selected_main == "Dock Leveler":
                st.markdown("##### 🏗️ Dock Leveler Specifications (Line-by-Line Order)")
                
                # 1st - CE Certification Details
                item["dock_ce_cert"] = st.multiselect(
                    "1st - CE Certification Details",
                    st.session_state["dock_ce_cert_list"],
                    default=item.get("dock_ce_cert", []),
                    key=f"dk_ce_{idx}"
                )
                
                # 2nd - Load Capacity
                item["dock_capacity"] = st.multiselect(
                    "2nd - Load Capacity",
                    st.session_state["dock_capacity_list"],
                    default=item.get("dock_capacity", []),
                    key=f"dk_cap_{idx}"
                )
                
                # 3rd - Cylinder
                item["dock_cylinder"] = st.multiselect(
                    "3rd - Cylinder",
                    st.session_state["dock_cylinder_list"],
                    default=item.get("dock_cylinder", []),
                    key=f"dk_cyl_{idx}"
                )
                
                # 4th - Height Adjustment of Dock Operation
                item["dock_height_adj"] = st.multiselect(
                    "4th - Height Adjustment of Dock Operation",
                    st.session_state["dock_height_adj_list"],
                    default=item.get("dock_height_adj", []),
                    key=f"dk_hgt_{idx}"
                )
                
                # 5th - Bumper
                item["dock_bumper"] = st.multiselect(
                    "5th - Bumper",
                    st.session_state["dock_bumper_list"],
                    default=item.get("dock_bumper", []),
                    key=f"dk_bmp_{idx}"
                )

            # --- ROLLING SHUTTER UI FIELDS ---
            elif selected_main == "Rolling Shutters":
                st.markdown("##### 📐 Shutter Technical Details")
                col_sp, col_sn = st.columns(2)
                with col_sp:
                    item["slat_pow"] = st.multiselect("Paint Finish", st.session_state["slat_pow_list"], default=item.get("slat_pow", []), key=f"sp_{idx}")
                with col_sn:
                    item["slat_nat"] = st.multiselect("Slats", st.session_state["slat_nat_list"], default=item.get("slat_nat", []), key=f"sn_{idx}")

                cg, cb, ch_col = st.columns(3)
                with cg:
                    item["guide"] = st.multiselect("Guide Specification", st.session_state["guide_list"], default=item.get("guide", []), key=f"gd_{idx}")
                with cb:
                    item["bottom"] = st.multiselect("Bottom Specification", st.session_state["bottom_list"], default=item.get("bottom", []), key=f"bt_{idx}")
                with ch_col:
                    item["hood"] = st.multiselect("Hood Cover Specification", st.session_state["hood_list"], default=item.get("hood", []), key=f"hd_{idx}")

                item["safety_locks"] = st.multiselect("Select Locks & Safety Features", SAFETY_LOCK_OPTIONS, default=item.get("safety_locks", []), key=f"lock_{idx}")
                if item["type"] == "Motorized Rolling Shutter":
                    item["operator"] = st.multiselect("Operator Option", OPERATOR_OPTIONS, default=item.get("operator", []), key=f"op_{idx}")

            st.markdown("---")
            st.markdown("**Dimensions & Rates:**")
            cw, ch, cq, cmr, cir = st.columns(5)
            item["width"] = cw.number_input("Width (mm)", value=int(item.get("width", 2000)), step=50, key=f"w_{idx}")
            item["height"] = ch.number_input("Length/Pit Depth (mm)", value=int(item.get("height", 2500)), step=50, key=f"h_{idx}")
            item["qty"] = cq.number_input("Qty", value=int(item.get("qty", 1)), min_value=1, step=1, key=f"q_{idx}")
            item["mat_rate"] = cmr.number_input("Material Rate (INR)", value=int(item.get("mat_rate", 245000)), step=500, key=f"mr_{idx}")
            item["inst_rate"] = cir.number_input("Installation Rate (INR)", value=int(item.get("inst_rate", 12000)), step=100, key=f"ir_{idx}")

    st.button("➕ Add Another Product Item", on_click=add_product_item)

    # --- EXTRA CHARGES ---
    st.markdown("### 🚚 Extra Charges & Expenses")
    total_mat_sum = sum(itm.get("qty", 1) * itm.get("mat_rate", 0) for itm in st.session_state["shutter_items"])

    col_pack_pct, col_pack_amt, c2, c3, c4, c5 = st.columns([1.2, 1.5, 2, 1.5, 1.5, 1.5])
    packing_pct = col_pack_pct.number_input("Packing & Loading (%)", value=2.5, step=0.5, key="pack_pct")
    packing_charges = int(round((total_mat_sum * packing_pct) / 100.0))
    col_pack_amt.text_input("Packing Amount (INR)", value=f"₹ {packing_charges:,}", disabled=True, key="pack_amt_disp")
    
    show_custom_freight = c2.checkbox("Hide Amount & Show Text", key="custom_freight_toggle")
    if show_custom_freight:
        freight_text_display = c2.text_input("Freight Display Text", value="Extra Charges", key="freight_text_input")
        freight_charges = 0
    else:
        freight_charges = c2.number_input("Freight Charges (INR)", value=18000)
        freight_text_display = f"{freight_charges:,}"

    unloading_charges = c3.number_input("Unloading Charges (INR)", value=0)
    crane_charges = c4.number_input("Crane Charges (INR)", value=0)
    scaffolding_charges = c5.number_input("Scaffolding Charges (INR)", value=0)

    # --- PDF GENERATOR ---
    def generate_pdf():
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=25,
            rightMargin=25,
            topMargin=20,
            bottomMargin=20,
        )
        story = []
        styles = getSampleStyleSheet()

        def get_header_element():
            logo_path = None
            for possible_path in ["Logo.jpeg", "logo.jpeg", "Logo.png", "logo.png"]:
                if os.path.exists(possible_path):
                    logo_path = possible_path
                    break
                    
            if logo_path:
                try:
                    logo_element = RLImage(logo_path, width=170, height=55)
                except Exception:
                    logo_element = Paragraph("<b>SIDHARTH</b><br/><font size=8 color='#003366'>SHUTTER & AUTOMATION</font>", styles["Normal"])
            else:
                logo_element = Paragraph("<b>SIDHARTH</b><br/><font size=8 color='#003366'>SHUTTER & AUTOMATION</font>", styles["Normal"])

            god_logo_path = None
            for possible_path in ["Logo2.jpeg", "logo2.jpeg", "Logo2.png", "logo2.png"]:
                if os.path.exists(possible_path):
                    god_logo_path = possible_path
                    break

            if god_logo_path:
                try:
                    god_logo_element = RLImage(god_logo_path, width=22, height=30)
                except Exception:
                    god_logo_element = Paragraph("", styles["Normal"])
            else:
                god_logo_element = Paragraph("", styles["Normal"])

            right_bold = ParagraphStyle("HeadRightBold", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9.5, leading=12, alignment=0)
            right_text = ParagraphStyle("HeadRightText", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=11, alignment=0, textColor=colors.HexColor("#1D4ED8"))

            right_table_data = [
                [Paragraph("<b>GSTIN/UIN: 08AEEPJ6848R1ZN</b>", right_bold)],
                [Paragraph("<b>Web:</b> <font color='#0284C7'>www.ssaapl.com</font>", right_text)],
                [Paragraph("<b>Email:</b> <font color='#0284C7'>sales@ssaapl.com</font>", right_text)],
                [Paragraph("<b>Ph:</b> <font color='#0284C7'>+91 90019 96526, +91 90010 42908</font>", right_text)],
                [Paragraph("<b>Add:</b> <font color='#0284C7'>H-1-89, RIICO Ind. Area, Mansarovar, Jaipur, Rajasthan, 302020</font>", right_text)],
            ]

            t_right_info = Table(right_table_data, colWidths=[230])
            t_right_info.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]))

            t_head = Table([[logo_element, god_logo_element, t_right_info]], colWidths=[185, 130, 230])
            t_head.setStyle(TableStyle([
                ("VALIGN", (0, 0), (0, 0), "MIDDLE"),
                ("VALIGN", (1, 0), (1, 0), "TOP"),
                ("VALIGN", (2, 0), (2, 0), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ]))
            return t_head

        style_cover_meta = ParagraphStyle("CoverMeta", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=15)
        style_cover_body = ParagraphStyle("CoverBody", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=15.5)
        style_sec_title = ParagraphStyle("SecTitle", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=12, leading=16)

        # PAGE 1: COVER LETTER
        story.append(get_header_element())
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#000000"), spaceAfter=15))

        meta_text = f"<b>Ref No:</b> {qtn_ref_no} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <b>Date:</b> {qtn_date}"
        story.append(Paragraph(meta_text, style_cover_meta))
        story.append(Spacer(1, 15))

        ship_formatted = shipping_address.replace("\n", "<br/>")
        client_info_p1 = (
            f"<b>To,</b><br/>"
            f"<b>{client_name}</b><br/>"
            f"{ship_formatted}<br/>"
            f"<b>Kind Attn:</b> {attn_person}<br/>"
            f"<b>Contact:</b> {contact_details}"
        )
        story.append(Paragraph(client_info_p1, style_cover_meta))
        story.append(Spacer(1, 20))

        body_formatted = cover_body_text.replace("\n", "<br/>")
        story.append(Paragraph(body_formatted, style_cover_body))
        story.append(Spacer(1, 35))

        sign_off_p1 = (
            "<b>For SIDHARTH SHUTTER & AUTOMATION PVT. LTD.</b><br/><br/><br/>"
            f"<b>Authorized Signatory</b><br/>"
            f"<b>({quotation_made_by})</b>"
        )
        story.append(Paragraph(sign_off_p1, style_cover_meta))

        # PAGE 2: ABOUT US
        if include_page_2:
            story.append(PageBreak())
            story.append(get_header_element())
            story.append(Spacer(1, 4))
            story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#000000"), spaceAfter=6))
            story.append(Paragraph("Section A: About Us & Our Expertise (Company Profile)", style_sec_title))
            story.append(Spacer(1, 6))

            style_box_head = ParagraphStyle("BoxHeadFull", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, leading=13)
            style_box_text = ParagraphStyle("BoxTextFull", parent=styles["Normal"], fontName="Helvetica", fontSize=9.5, leading=13)

            about_box_content = []
            for ab_item in st.session_state["about_us_data"]:
                about_box_content.append(Paragraph(f"<b>{ab_item['title']}</b>", style_box_head))
                about_box_content.append(Spacer(1, 2))
                for line in ab_item['text'].split("\n"):
                    if line.strip():
                        about_box_content.append(Paragraph(line.replace("•", "&bull;"), style_box_text))
                        about_box_content.append(Spacer(1, 2))

            t_box = Table([[about_box_content]], colWidths=[545])
            t_box.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#000000")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(t_box)

        # PAGE 3: COMMERCIAL TABLE
        story.append(PageBreak())
        story.append(get_header_element())
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#000000"), spaceAfter=8))

        small_bold = ParagraphStyle("SmallBold", fontName="Helvetica-Bold", fontSize=8, leading=10)
        small_bold_center = ParagraphStyle("SmallBoldCenter", fontName="Helvetica-Bold", fontSize=8, leading=10, alignment=1)
        small_bold_right = ParagraphStyle("SmallBoldRight", fontName="Helvetica-Bold", fontSize=8, leading=10, alignment=2)
        small_text = ParagraphStyle("SmallText", fontName="Helvetica", fontSize=8, leading=10)
        small_text_center = ParagraphStyle("SmallTextCenter", fontName="Helvetica", fontSize=8, leading=10, alignment=1)
        small_text_right = ParagraphStyle("SmallTextRight", fontName="Helvetica", fontSize=8, leading=10, alignment=2)

        header_data = [
            [Paragraph("<b>Company Name</b>", small_bold), Paragraph(client_name, small_text), Paragraph("<b>Client ID:</b>", small_bold), Paragraph(client_id, small_bold)],
            [Paragraph("<b>Contact Details</b>", small_bold), Paragraph(contact_details, small_text), Paragraph("<b>Qtn Date:</b>", small_bold), Paragraph(str(qtn_date), small_text)],
            [Paragraph("<b>Shipping Address</b>", small_bold), Paragraph(shipping_address, small_text), Paragraph("<b>Qtn Ref No:</b>", small_bold), Paragraph(qtn_ref_no, small_bold)],
            [Paragraph("<b>Billing Address</b>", small_bold), Paragraph(billing_address, small_text), Paragraph("<b>Sales Reference:</b>", small_bold), Paragraph(f"{sales_reference_1}<br/>{sales_reference_2}", small_text)],
            [Paragraph("<b>GSTIN</b>", small_bold), Paragraph(gstin, small_text), Paragraph("<b>Site Contact:</b>", small_bold), Paragraph(site_person, small_text)],
        ]

        t_header = Table(header_data, colWidths=[90, 230, 85, 140])
        t_header.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#000000")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1E293B")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t_header)
        story.append(Spacer(1, 10))

        items_data = [[
            Paragraph("<b>Sr. No.</b>", small_bold_center),
            Paragraph("<b>Description</b>", small_bold),
            Paragraph("<b>HSN Code</b>", small_bold_center),
            Paragraph("<b>Width (mm)</b>", small_bold_center),
            Paragraph("<b>Height/Len (mm)</b>", small_bold_center),
            Paragraph("<b>Qty</b>", small_bold_center),
            Paragraph("<b>Mat. Rate (INR)</b>", small_bold_right),
            Paragraph("<b>Mat. Amount (INR)</b>", small_bold_right),
            Paragraph("<b>Inst. Rate (INR)</b>", small_bold_right),
            Paragraph("<b>Inst. Amount (INR)</b>", small_bold_right),
        ]]

        mat_grand_total = 0
        inst_grand_total = 0
        total_qty = 0

        for i, itm in enumerate(st.session_state["shutter_items"]):
            m_amt = itm.get("qty", 1) * itm.get("mat_rate", 0)
            i_amt = itm.get("qty", 1) * itm.get("inst_rate", 0)
            mat_grand_total += m_amt
            inst_grand_total += i_amt
            total_qty += itm.get("qty", 1)

            desc_lines = [f"<b>{itm.get('type', 'Hydraulic Doclevller')}</b>"]
            
            # UPDATED DOCK LEVELER LINE BY LINE DESCRIPTION IN PDF
            if itm.get("main_product") == "Dock Leveler":
                for ce in itm.get("dock_ce_cert", []):
                    desc_lines.append(f"- Certification: {ce}")
                for cap in itm.get("dock_capacity", []):
                    desc_lines.append(f"- Capacity: {cap}")
                for cyl in itm.get("dock_cylinder", []):
                    desc_lines.append(f"- Cylinder: {cyl}")
                for hgt in itm.get("dock_height_adj", []):
                    desc_lines.append(f"- Height Adj.: {hgt}")
                for bmp in itm.get("dock_bumper", []):
                    desc_lines.append(f"- Bumper: {bmp}")
                for sft in itm.get("dock_safety", []):
                    desc_lines.append(f"- {sft}")
            else:
                for s in itm.get("slat_pow", []):
                    desc_lines.append(f"- {s}")
                for s in itm.get("slat_nat", []):
                    desc_lines.append(f"- {s}")
                for g in itm.get("guide", []):
                    desc_lines.append(f"- {g}")
                for b in itm.get("bottom", []):
                    desc_lines.append(f"- {b}")
                for h in itm.get("hood", []):
                    desc_lines.append(f"- {h}")
                for l in itm.get("safety_locks", []):
                    desc_lines.append(f"- {l}")

            desc = "<br/>".join(desc_lines)

            items_data.append([
                Paragraph(str(i + 1), small_text_center),
                Paragraph(desc, small_text),
                Paragraph(itm.get("hsn", "-"), small_text_center),
                Paragraph(str(itm.get("width", "-")), small_text_center),
                Paragraph(str(itm.get("height", "-")), small_text_center),
                Paragraph(str(itm.get("qty", "-")), small_text_center),
                Paragraph(f"{itm.get('mat_rate', 0):,}", small_text_right),
                Paragraph(f"{m_amt:,}", small_text_right),
                Paragraph(f"{itm.get('inst_rate', 0):,}", small_text_right),
                Paragraph(f"{i_amt:,}", small_text_right),
            ])

        subtotal_extras = packing_charges + freight_charges + unloading_charges + crane_charges + scaffolding_charges
        subtotal_mat = mat_grand_total + subtotal_extras
        gst_mat = round(subtotal_mat * 0.18)
        gst_inst = round(inst_grand_total * 0.18)

        total_mat_with_gst = subtotal_mat + gst_mat
        total_inst_with_gst = inst_grand_total + gst_inst
        final_grand_total = total_mat_with_gst + total_inst_with_gst

        items_data.append(["", Paragraph("<b>Item Total</b>", small_bold), "", "", "", Paragraph(f"<b>{total_qty}</b>", small_bold_center), "", Paragraph(f"<b>{mat_grand_total:,}</b>", small_bold_right), "", Paragraph(f"<b>{inst_grand_total:,}</b>", small_bold_right)])
        items_data.append(["", Paragraph(f"Packing Charges ({packing_pct}%)", small_text), "", "", "", "", "", Paragraph(f"{packing_charges:,}", small_text_right), "", Paragraph("-", small_text_right)])
        items_data.append(["", Paragraph("Freight Charges (As Per Actual)", small_text), "", "", "", "", "", Paragraph(freight_text_display, small_text_right), "", Paragraph("-", small_text_right)])

        items_data.append(["", Paragraph("<b>Supply & Installation Amount Excluding GST</b>", small_bold), "", "", "", "", "", Paragraph(f"<b>{subtotal_mat:,}</b>", small_bold_right), "", Paragraph(f"<b>{inst_grand_total:,}</b>", small_bold_right)])
        items_data.append(["", Paragraph("GST on supply & installation @18%", small_text), "", "", "", "", "", Paragraph(f"{gst_mat:,}", small_text_right), Paragraph("@18%", small_text_right), Paragraph(f"{gst_inst:,}", small_text_right)])
        items_data.append(["", Paragraph("<b>Total supply & installation with GST</b>", small_bold), "", "", "", "", "", Paragraph(f"<b>{total_mat_with_gst:,}</b>", small_bold_right), "", Paragraph(f"<b>{total_inst_with_gst:,}</b>", small_bold_right)])
        items_data.append(["", Paragraph("<b>Grand total with GST</b>", small_bold), "", "", "", "", "", "", "", Paragraph(f"<b>{final_grand_total:,}</b>", small_bold_right)])

        t_items = Table(items_data, colWidths=[22, 160, 50, 38, 38, 22, 55, 60, 50, 60])
        t_items.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t_items)

        # PAGE 4: TECHNICAL SPECS
        if include_page_4:
            story.append(PageBreak())
            story.append(get_header_element())
            story.append(Spacer(1, 6))
            story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#000000"), spaceAfter=8))
            story.append(Paragraph("Section C: Technical Specifications", style_sec_title))
            story.append(Spacer(1, 4))

            style_th_param = ParagraphStyle("ThParam", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9.5)
            style_th_spec = ParagraphStyle("ThSpec", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9.5, alignment=1)
            style_td_param = ParagraphStyle("TdParam", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9)
            style_td_spec = ParagraphStyle("TdSpec", parent=styles["Normal"], fontName="Helvetica", fontSize=9)

            tech_table_data = [[Paragraph("Parameter", style_th_param), Paragraph("Specification Details", style_th_spec)]]
            for spec_item in st.session_state["tech_specs_data"]:
                tech_table_data.append([Paragraph(spec_item["param"], style_td_param), Paragraph(spec_item["spec"], style_td_spec)])

            t_tech = Table(tech_table_data, colWidths=[140, 405])
            t_tech.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#000000")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#000000")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(t_tech)

        # PAGE 5: TERMS & CONDITIONS
        if include_page_5:
            story.append(PageBreak())
            story.append(get_header_element())
            story.append(Spacer(1, 6))
            story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#000000"), spaceAfter=8))
            story.append(Paragraph("Section D: Annexure – Terms & Condition", style_sec_title))
            story.append(Spacer(1, 4))

            style_tc_cat = ParagraphStyle("TcCat", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9)
            style_tc_det = ParagraphStyle("TcDet", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5)

            terms_table_data = []
            for term in st.session_state["terms_data"]:
                details_formatted = term["details"].replace("\n", "<br/>").replace("•", "&bull;")
                terms_table_data.append([Paragraph(f"<b>{term['category']}</b>", style_tc_cat), Paragraph(details_formatted, style_tc_det)])

            t_terms = Table(terms_table_data, colWidths=[135, 410])
            t_terms.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#000000")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#000000")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(t_terms)

        doc.build(story)
        buffer.seek(0)
        return buffer

    # EXCEL GENERATOR
    def generate_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotation"

        font_company = Font(name="Calibri", size=14, bold=True)
        font_header_bold = Font(name="Calibri", size=10, bold=True)
        font_regular = Font(name="Calibri", size=9)

        fill_table_header = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws.merge_cells("A1:J1")
        ws["A1"] = "SIDHARTH SHUTTER & AUTOMATION PRIVATE LIMITED"
        ws["A1"].font = font_company
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Sr. No.", "Description", "HSN Code", "Width (mm)", "Height (mm)", "Qty", "Mat. Rate (INR)", "Mat. Amount (INR)", "Inst. Rate (INR)", "Inst. Amount (INR)"]
        curr_row = 4
        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=curr_row, column=col_num, value=h_text)
            cell.font = font_header_bold
            cell.fill = fill_table_header
            cell.border = thin_border

        curr_row += 1
        for i, itm in enumerate(st.session_state["shutter_items"]):
            m_amt = itm.get("qty", 1) * itm.get("mat_rate", 0)
            i_amt = itm.get("qty", 1) * itm.get("inst_rate", 0)

            desc_lines = [itm.get("type", "Hydraulic Doclevller")]
            if itm.get("main_product") == "Dock Leveler":
                for ce in itm.get("dock_ce_cert", []): desc_lines.append(f"• Certification: {ce}")
                for cap in itm.get("dock_capacity", []): desc_lines.append(f"• Capacity: {cap}")
                for cyl in itm.get("dock_cylinder", []): desc_lines.append(f"• Cylinder: {cyl}")
                for hgt in itm.get("dock_height_adj", []): desc_lines.append(f"• Height Adj.: {hgt}")
                for bmp in itm.get("dock_bumper", []): desc_lines.append(f"• Bumper: {bmp}")
            else:
                for s in itm.get("slat_pow", []): desc_lines.append(f"• {s}")
                for s in itm.get("slat_nat", []): desc_lines.append(f"• {s}")

            row_data = [i + 1, "\n".join(desc_lines), itm.get("hsn", "-"), itm.get("width", "-"), itm.get("height", "-"), itm.get("qty", "-"), itm.get("mat_rate", 0), m_amt, itm.get("inst_rate", 0), i_amt]
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=col_num, value=val)
                cell.font = font_regular
                cell.border = thin_border

            curr_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- ACTIONS & PREVIEW ---
    st.markdown("---")
    st.markdown("### 📥 Generate & Export Proposals")

    col_pdf, col_excel = st.columns(2)
    with col_pdf:
        if st.button("👁️ Preview PDF Quotation", type="primary", use_container_width=True):
            st.session_state["pdf_preview_bytes"] = generate_pdf().getvalue()

    with col_excel:
        if st.button("📊 Generate Excel Quotation", use_container_width=True):
            st.download_button(
                label="📥 Download Quotation Excel (.xlsx)",
                data=generate_excel(),
                file_name=f"Quotation_{qtn_ref_no.replace('/', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if "pdf_preview_bytes" in st.session_state:
        st.markdown("---")
        st.markdown("### 🔍 PDF Quotation Preview & Download")
        pdf_data = st.session_state["pdf_preview_bytes"]

        c_dl, c_cls = st.columns([2, 1])
        with c_dl:
            st.download_button("📥 Download Verified Multi-Page PDF Quotation", data=pdf_data, file_name=f"Quotation_{qtn_ref_no.replace('/', '_')}.pdf", mime="application/pdf", type="primary", use_container_width=True)
        with c_cls:
            if st.button("❌ Close Preview", use_container_width=True):
                del st.session_state["pdf_preview_bytes"]
                st.rerun()

        pdf = pdfium.PdfDocument(pdf_data)
        for i, page in enumerate(pdf):
            image = page.render(scale=2).to_pil()
            st.image(image, caption=f"Page {i+1}", use_container_width=True)

else:
    st.info(f"⏳ **{st.session_state['selected_product']} Quotation Builder** is currently under development.")
